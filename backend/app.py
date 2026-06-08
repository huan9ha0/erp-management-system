#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
企业进销存智能管理系统 - 核心功能后端API
模块：库存管理、进货管理、销售管理
技术栈：Flask + SQLAlchemy + SQLite
"""

import os
import sys
import csv
import io
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, case, text, or_
from sqlalchemy.exc import IntegrityError
from contextlib import contextmanager

# ============================================================
# 应用初始化
# ============================================================
app = Flask(__name__, static_folder='../frontend', static_url_path='')
CORS(app)

# 全局错误处理：API路由返回JSON而非HTML
@app.errorhandler(500)
def internal_error(e):
    return jsonify({'error': f'服务器内部错误: {str(e)}'}), 500

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': '接口不存在'}), 404

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(BASE_DIR, "erp.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


def paginate_query(query, page, per_page):
    """对SQLAlchemy query对象进行分页，返回分页结果字典"""
    if page < 1:
        page = 1
    total = query.count()
    total_pages = max((total + per_page - 1) // per_page, 1) if total > 0 else 1
    items = query.offset((page - 1) * per_page).limit(per_page).all()
    return {
        'data': [item.to_dict() for item in items],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages
    }


# ============================================================
# 数据库模型设计
# ============================================================

class Supplier(db.Model):
    """供应商表"""
    __tablename__ = 'suppliers'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(100), nullable=False, comment='供应商名称')
    contact_person = db.Column(db.String(50), comment='联系人')
    phone = db.Column(db.String(20), comment='联系电话')
    email = db.Column(db.String(100), comment='邮箱')
    address = db.Column(db.String(200), comment='地址')
    rating = db.Column(db.Integer, default=3, comment='评级(1-5)')
    cooperation_history = db.Column(db.Text, comment='合作历史')
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id, 'name': self.name,
            'contact_person': self.contact_person, 'phone': self.phone,
            'email': self.email, 'address': self.address,
            'rating': self.rating, 'cooperation_history': self.cooperation_history,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else ''
        }


class Customer(db.Model):
    """客户表"""
    __tablename__ = 'customers'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    name = db.Column(db.String(100), nullable=False, comment='客户名称')
    contact_person = db.Column(db.String(50), comment='联系人')
    phone = db.Column(db.String(20), comment='联系电话')
    email = db.Column(db.String(100), comment='邮箱')
    address = db.Column(db.String(200), comment='地址')
    region = db.Column(db.String(50), comment='所在区域')
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id, 'name': self.name,
            'contact_person': self.contact_person, 'phone': self.phone,
            'email': self.email, 'address': self.address,
            'region': self.region,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else ''
        }


class Product(db.Model):
    """商品表"""
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    code = db.Column(db.String(50), unique=True, nullable=False, comment='商品编码')
    name = db.Column(db.String(100), nullable=False, comment='商品名称')
    category = db.Column(db.String(50), comment='商品类别')
    model = db.Column(db.String(100), comment='型号')
    manufacturer = db.Column(db.String(100), comment='生产厂商')
    unit = db.Column(db.String(20), default='个', comment='单位')
    purchase_price = db.Column(db.Float, default=0, comment='采购参考价')
    sale_price = db.Column(db.Float, default=0, comment='销售单价')
    safety_stock = db.Column(db.Integer, default=10, comment='安全库存量')
    created_at = db.Column(db.DateTime, default=datetime.now)

    # 关联实时库存
    inventory = db.relationship('Inventory', backref='product', uselist=False, lazy='joined')

    def to_dict(self):
        inv = self.inventory
        current_stock = inv.quantity if inv else 0
        return {
            'id': self.id, 'code': self.code, 'name': self.name,
            'category': self.category, 'model': self.model,
            'manufacturer': self.manufacturer, 'unit': self.unit,
            'purchase_price': self.purchase_price, 'sale_price': self.sale_price,
            'safety_stock': self.safety_stock,
            'current_stock': current_stock,
            'inventory_value': round(current_stock * self.purchase_price, 2),
            'stock_status': self._stock_status(inv),
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else ''
        }

    def _stock_status(self, inv):
        if not inv or inv.quantity == 0:
            return '缺货'
        if inv.quantity <= self.safety_stock:
            return '预警'
        if inv.quantity <= self.safety_stock * 2:
            return '偏低'
        return '正常'


class Inventory(db.Model):
    """实时库存表"""
    __tablename__ = 'inventory'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), unique=True, nullable=False)
    quantity = db.Column(db.Integer, default=0, comment='当前库存数量')
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id, 'product_id': self.product_id,
            'quantity': self.quantity,
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class InventoryLog(db.Model):
    """库存变动日志"""
    __tablename__ = 'inventory_logs'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    change_type = db.Column(db.String(20), comment='变动类型: purchase_in/sales_out/adjust/inventory_check')
    change_quantity = db.Column(db.Integer, nullable=False, comment='变动数量(正=入库,负=出库)')
    before_quantity = db.Column(db.Integer, comment='变动前库存')
    after_quantity = db.Column(db.Integer, comment='变动后库存')
    reference_no = db.Column(db.String(50), comment='关联单号')
    remark = db.Column(db.String(200), comment='备注')
    created_at = db.Column(db.DateTime, default=datetime.now)

    product = db.relationship('Product', backref='inventory_logs')

    def to_dict(self):
        return {
            'id': self.id, 'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'change_type': self.change_type, 'change_quantity': self.change_quantity,
            'before_quantity': self.before_quantity, 'after_quantity': self.after_quantity,
            'reference_no': self.reference_no, 'remark': self.remark,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


class PurchaseOrder(db.Model):
    """采购订单表"""
    __tablename__ = 'purchase_orders'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_no = db.Column(db.String(50), unique=True, nullable=False, comment='订单编号')
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'), nullable=False)
    status = db.Column(db.String(20), default='draft', comment='状态: draft/approved/paid/received/cancelled')
    total_amount = db.Column(db.Float, default=0, comment='总金额')
    remark = db.Column(db.String(200), comment='备注')
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    supplier = db.relationship('Supplier', backref='purchase_orders')
    items = db.relationship('PurchaseOrderItem', backref='order', lazy='joined',
                            cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id, 'order_no': self.order_no,
            'supplier_id': self.supplier_id,
            'supplier_name': self.supplier.name if self.supplier else '',
            'status': self.status, 'total_amount': self.total_amount,
            'remark': self.remark,
            'items': [item.to_dict() for item in self.items],
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else '',
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M') if self.updated_at else ''
        }


class PurchaseOrderItem(db.Model):
    """采购订单明细表"""
    __tablename__ = 'purchase_order_items'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_id = db.Column(db.Integer, db.ForeignKey('purchase_orders.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, comment='采购数量')
    unit_price = db.Column(db.Float, nullable=False, comment='采购单价')
    subtotal = db.Column(db.Float, comment='小计金额')

    product = db.relationship('Product')

    def to_dict(self):
        return {
            'id': self.id, 'order_id': self.order_id,
            'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'quantity': self.quantity, 'unit_price': self.unit_price,
            'subtotal': self.subtotal or (self.quantity * self.unit_price)
        }


class SalesOrder(db.Model):
    """销售订单表"""
    __tablename__ = 'sales_orders'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_no = db.Column(db.String(50), unique=True, nullable=False, comment='订单编号')
    customer_id = db.Column(db.Integer, db.ForeignKey('customers.id'), nullable=False)
    status = db.Column(db.String(20), default='draft', comment='状态: draft/confirmed/collected/shipped/completed/cancelled')
    total_amount = db.Column(db.Float, default=0, comment='总金额')
    remark = db.Column(db.String(200), comment='备注')
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    customer = db.relationship('Customer', backref='sales_orders')
    items = db.relationship('SalesOrderItem', backref='order', lazy='joined',
                            cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id, 'order_no': self.order_no,
            'customer_id': self.customer_id,
            'customer_name': self.customer.name if self.customer else '',
            'status': self.status, 'total_amount': self.total_amount,
            'remark': self.remark,
            'items': [item.to_dict() for item in self.items],
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else '',
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M') if self.updated_at else ''
        }


class SalesOrderItem(db.Model):
    """销售订单明细表"""
    __tablename__ = 'sales_order_items'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_id = db.Column(db.Integer, db.ForeignKey('sales_orders.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, comment='销售数量')
    unit_price = db.Column(db.Float, nullable=False, comment='销售单价')
    subtotal = db.Column(db.Float, comment='小计金额')

    product = db.relationship('Product')

    def to_dict(self):
        return {
            'id': self.id, 'order_id': self.order_id,
            'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'quantity': self.quantity, 'unit_price': self.unit_price,
            'subtotal': self.subtotal or (self.quantity * self.unit_price)
        }


class ReturnOrder(db.Model):
    """退货单主表"""
    __tablename__ = 'return_orders'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_no = db.Column(db.String(30), unique=True, nullable=False, comment='退货单号')
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'), nullable=False, comment='供应商ID')
    status = db.Column(db.String(20), default='draft', comment='draft/submitted/approved/returned/refunded/completed/cancelled')
    total_amount = db.Column(db.Float, default=0, comment='退货总金额')
    remark = db.Column(db.String(200), comment='备注')
    created_at = db.Column(db.DateTime, default=datetime.now, comment='创建时间')
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now, comment='更新时间')

    supplier = db.relationship('Supplier')
    items = db.relationship('ReturnOrderItem', backref='order', lazy='dynamic', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id, 'order_no': self.order_no,
            'supplier_id': self.supplier_id,
            'supplier_name': self.supplier.name if self.supplier else '',
            'status': self.status, 'total_amount': self.total_amount,
            'remark': self.remark,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else '',
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else '',
            'items': [item.to_dict() for item in self.items.all()]
        }


class ReturnOrderItem(db.Model):
    """退货单明细表"""
    __tablename__ = 'return_order_items'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_id = db.Column(db.Integer, db.ForeignKey('return_orders.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, comment='退货数量')
    unit_price = db.Column(db.Float, nullable=False, comment='退货单价')
    subtotal = db.Column(db.Float, comment='小计金额')

    product = db.relationship('Product')

    def to_dict(self):
        return {
            'id': self.id, 'order_id': self.order_id,
            'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'quantity': self.quantity, 'unit_price': self.unit_price,
            'subtotal': self.subtotal or (self.quantity * self.unit_price)
        }


class CustomerReturnOrder(db.Model):
    """客户退货单主表"""
    __tablename__ = 'customer_return_orders'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_no = db.Column(db.String(30), unique=True, nullable=False, comment='退货单号')
    customer_id = db.Column(db.Integer, db.ForeignKey('customers.id'), nullable=False, comment='客户ID')
    status = db.Column(db.String(20), default='draft', comment='draft/submitted/approved/received/refunded/completed/cancelled')
    total_amount = db.Column(db.Float, default=0, comment='退货总金额')
    remark = db.Column(db.String(200), comment='备注')
    created_at = db.Column(db.DateTime, default=datetime.now, comment='创建时间')
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now, comment='更新时间')

    customer = db.relationship('Customer')
    items = db.relationship('CustomerReturnOrderItem', backref='order', lazy='dynamic', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id, 'order_no': self.order_no,
            'customer_id': self.customer_id,
            'customer_name': self.customer.name if self.customer else '',
            'status': self.status, 'total_amount': self.total_amount,
            'remark': self.remark,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else '',
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else '',
            'items': [item.to_dict() for item in self.items.all()]
        }


class CustomerReturnOrderItem(db.Model):
    """客户退货单明细表"""
    __tablename__ = 'customer_return_order_items'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    order_id = db.Column(db.Integer, db.ForeignKey('customer_return_orders.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, comment='退货数量')
    unit_price = db.Column(db.Float, nullable=False, comment='退货单价')
    subtotal = db.Column(db.Float, comment='小计金额')

    product = db.relationship('Product')

    def to_dict(self):
        return {
            'id': self.id, 'order_id': self.order_id,
            'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'quantity': self.quantity, 'unit_price': self.unit_price,
            'subtotal': self.subtotal or (self.quantity * self.unit_price)
        }


# ============================================================
# 工具函数
# ============================================================

def generate_order_no(prefix):
    """生成订单编号: 前缀 + 年月日 + 序号"""
    today = datetime.now().strftime('%Y%m%d')
    return f'{prefix}{today}{datetime.now().strftime("%H%M%S%f")[:-3]}'


@contextmanager
def transaction():
    """事务管理"""
    try:
        yield
        db.session.commit()
    except Exception as e:
        db.session.flush()
        db.session.rollback()
        raise e


def update_inventory(product_id, change_qty, change_type, ref_no, remark=''):
    """
    核心：更新库存（自动同步）
    - 采购入库: change_qty > 0
    - 销售出库: change_qty < 0
    返回: (成功标志, 消息)
    """
    product = Product.query.get(product_id)
    if not product:
        return False, '商品不存在'

    inv = Inventory.query.filter_by(product_id=product_id).first()
    if not inv:
        inv = Inventory(product_id=product_id, quantity=0)
        db.session.add(inv)
        db.session.flush()

    before_qty = inv.quantity

    # 销售出库检查库存是否充足
    if change_qty < 0 and inv.quantity + change_qty < 0:
        return False, f'商品[{product.name}]库存不足！当前库存: {inv.quantity}, 需要: {abs(change_qty)}'

    inv.quantity += change_qty
    after_qty = inv.quantity

    # 记录库存变动日志
    log = InventoryLog(
        product_id=product_id,
        change_type=change_type,
        change_quantity=change_qty,
        before_quantity=before_qty,
        after_quantity=after_qty,
        reference_no=ref_no,
        remark=remark
    )
    db.session.add(log)

    return True, '库存更新成功'


def get_stock_alerts():
    """获取库存预警列表"""
    results = db.session.query(
        Product, Inventory
    ).outerjoin(Inventory, Product.id == Inventory.product_id).all()

    alerts = []
    for product, inv in results:
        qty = inv.quantity if inv else 0
        if qty <= product.safety_stock:
            alerts.append({
                'product_id': product.id,
                'product_name': product.name,
                'product_code': product.code,
                'current_stock': qty,
                'safety_stock': product.safety_stock,
                'gap': product.safety_stock - qty,
                'level': '严重' if qty == 0 else '预警'
            })
    return alerts


# ============================================================
# 路由 - 首页 & 静态文件
# ============================================================

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')


# ============================================================
# 路由 - 仪表盘API
# ============================================================

@app.route('/api/dashboard', methods=['GET'])
def dashboard():
    """仪表盘数据"""
    total_products = Product.query.count()
    total_suppliers = Supplier.query.count()
    total_customers = Customer.query.count()

    # 库存总价值
    total_inventory_value = db.session.query(
        func.coalesce(func.sum(Inventory.quantity * Product.purchase_price), 0)
    ).join(Product, Inventory.product_id == Product.id).scalar()

    # 本月采购总额
    month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_purchase = db.session.query(
        func.coalesce(func.sum(PurchaseOrder.total_amount), 0)
    ).filter(PurchaseOrder.status == 'received',
             PurchaseOrder.updated_at >= month_start).scalar()

    # 本月销售总额
    month_sales = db.session.query(
        func.coalesce(func.sum(SalesOrder.total_amount), 0)
    ).filter(SalesOrder.status == 'completed',
             SalesOrder.updated_at >= month_start).scalar()

    # 库存预警
    alerts = get_stock_alerts()

    # 最近采购订单
    recent_purchase = PurchaseOrder.query.order_by(
        PurchaseOrder.created_at.desc()).limit(5).all()

    # 最近销售订单
    recent_sales = SalesOrder.query.order_by(
        SalesOrder.created_at.desc()).limit(5).all()

    return jsonify({
        'summary': {
            'total_products': total_products,
            'total_suppliers': total_suppliers,
            'total_customers': total_customers,
            'total_inventory_value': round(total_inventory_value, 2),
            'month_purchase': round(month_purchase, 2),
            'month_sales': round(month_sales, 2),
            'alert_count': len(alerts)
        },
        'alerts': alerts,
        'recent_purchase': [o.to_dict() for o in recent_purchase],
        'recent_sales': [o.to_dict() for o in recent_sales]
    })


# ============================================================
# 路由 - 商品管理API
# ============================================================

@app.route('/api/products', methods=['GET'])
def list_products():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    query = Product.query.order_by(Product.id.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/products', methods=['POST'])
def create_product():
    data = request.json
    if not data.get('name'):
        return jsonify({'error': '商品名称不能为空'}), 400

    try:
        product = Product(
            code=data.get('code', ''),
            name=data['name'],
            category=data.get('category', ''),
            model=data.get('model', ''),
            manufacturer=data.get('manufacturer', ''),
            unit=data.get('unit', '个'),
            purchase_price=data.get('purchase_price', 0),
            sale_price=data.get('sale_price', 0),
            safety_stock=data.get('safety_stock', 10)
        )
        db.session.add(product)
        db.session.flush()

        # 自动初始化库存记录
        initial_stock = int(data.get('initial_stock', 0) or 0)
        inv = Inventory(product_id=product.id, quantity=initial_stock)
        db.session.add(inv)
        db.session.flush()

        # 记录库存初始化日志（无论库存是否为0都记录）
        log = InventoryLog(
            product_id=product.id, change_type='create',
            change_quantity=initial_stock, before_quantity=0,
            after_quantity=initial_stock,
            reference_no=f'NEW-{product.code}',
            remark='新建商品' + (f'，初始库存 {initial_stock}' if initial_stock > 0 else '，库存为0')
        )
        db.session.add(log)

        db.session.commit()

        return jsonify(product.to_dict()), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': f'商品编码 "{data.get("code", "")}" 已存在，请更换编码'}), 400
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'创建失败: {str(e)}'}), 500


@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    product = Product.query.get(pid)
    if not product:
        return jsonify({'error': '商品不存在'}), 404
    data = request.json
    for key in ['name', 'code', 'category', 'model', 'manufacturer', 'unit', 'purchase_price', 'sale_price', 'safety_stock']:
        if key in data:
            setattr(product, key, data[key])
    db.session.commit()
    return jsonify(product.to_dict())


@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    """删除商品（同时清理关联的库存和日志）"""
    product = Product.query.get(pid)
    if not product:
        return jsonify({'error': '商品不存在'}), 404

    try:
        with transaction():
            # 删除关联库存记录
            Inventory.query.filter_by(product_id=pid).delete()
            # 删除库存变动日志
            InventoryLog.query.filter_by(product_id=pid).delete()
            # 删除商品本身
            db.session.delete(product)
        return jsonify({'message': '删除成功'})
    except Exception as e:
        return jsonify({'error': f'删除失败: {str(e)}'}), 400


@app.route('/api/products/batch-delete', methods=['POST'])
def batch_delete_products():
    """批量删除商品（同时清理关联的库存和日志）"""
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的商品ID列表'}), 400

    try:
        with transaction():
            for pid in ids:
                product = Product.query.get(pid)
                if not product:
                    continue
                Inventory.query.filter_by(product_id=pid).delete()
                InventoryLog.query.filter_by(product_id=pid).delete()
                db.session.delete(product)
        return jsonify({'message': f'成功删除 {len(ids)} 个商品'})
    except Exception as e:
        return jsonify({'error': f'批量删除失败: {str(e)}'}), 400


@app.route('/api/products/import', methods=['POST'])
def import_products():
    """
    批量导入商品（CSV/Excel 文件）
    文件列：商品编码,商品名称,类别,型号,供应商,单位,采购价,销售价,安全库存,库存数量
    - 全新编码：创建商品 + 生成草稿采购订单
    - 重复编码：纳入草稿采购订单
    - 按供应商分组，每组生成一张采购订单
    - 订单保持 draft 状态，需人工操作审批→付款→入库
    """
    if 'file' not in request.files:
        return jsonify({'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    # 根据扩展名判断文件类型
    filename = file.filename.lower()

    try:
        # 读取文件内容
        content = file.read()

        if filename.endswith('.csv'):
            rows = _parse_csv(content)
        elif filename.endswith(('.xls', '.xlsx')):
            # 尝试用 openpyxl 解析，fallback 到 csv
            try:
                rows = _parse_xlsx(content)
            except ImportError:
                return jsonify({'error': 'Excel解析需要 openpyxl 库，请用 CSV 格式或安装: pip install openpyxl'}), 400
        else:
            return jsonify({'error': '不支持的文件格式，请上传 .csv 或 .xlsx 文件'}), 400

        created = 0
        updated = 0
        skipped = 0
        errors = []
        created_products = []
        processed_items = []  # 所有需要纳入采购订单的行: {product, quantity, unit_price, supplier_name}

        with transaction():
            for i, row in enumerate(rows, start=2):  # 第1行是表头
                code = (row.get('code') or row.get('商品编码') or '').strip()
                name = (row.get('name') or row.get('商品名称') or '').strip()

                if not code or not name:
                    errors.append(f'第{i}行: 商品编码或名称为空，跳过')
                    skipped += 1
                    continue

                # 读取数量和采购价
                _qty_str = (row.get('quantity') or row.get('库存数量') or '0').strip()
                try:
                    qty = int(float(_qty_str)) if _qty_str else 0
                except (ValueError, TypeError):
                    qty = 0

                _price_str = (row.get('purchase_price') or row.get('采购价') or '0').strip()
                try:
                    unit_price = float(_price_str) if _price_str else 0
                except (ValueError, TypeError):
                    unit_price = 0

                supplier_name = (row.get('supplier') or row.get('供应商')
                                 or row.get('manufacturer') or row.get('生产厂商') or '').strip()

                # 检查编码是否已存在
                existing = Product.query.filter_by(code=code).first()
                if existing:
                    # 已存在：纳入采购订单，库存由订单入库流程统一处理
                    if qty == 0:
                        errors.append(f'第{i}行({name}): 编码 {code} 已存在，但数量为0，未产生采购记录')
                        skipped += 1
                        continue

                    # 如果导入文件未提供供应商，沿用商品已有的 manufacturer
                    if not supplier_name:
                        supplier_name = existing.manufacturer or ''

                    processed_items.append({
                        'product': existing,
                        'quantity': qty,
                        'unit_price': unit_price or existing.purchase_price,
                        'supplier_name': supplier_name
                    })
                    errors.append(f'第{i}行({name}): 编码 {code} 已存在，库存 +{qty}，已纳入采购订单')
                    updated += 1
                    continue

                # 全新商品
                try:
                    product = Product(
                        code=code,
                        name=name,
                        category=(row.get('category') or row.get('分类') or '').strip(),
                        model=(row.get('model') or row.get('型号') or '').strip(),
                        manufacturer=supplier_name,
                        unit=(row.get('unit') or row.get('单位') or '个').strip(),
                        purchase_price=unit_price,
                        sale_price=float(row.get('sale_price') or row.get('销售价') or 0),
                        safety_stock=int(row.get('safety_stock') or row.get('安全库存') or 10)
                    )
                    db.session.add(product)
                    db.session.flush()

                    created_products.append(product.to_dict())
                    created += 1

                    processed_items.append({
                        'product': product,
                        'quantity': qty,
                        'unit_price': unit_price,
                        'supplier_name': supplier_name
                    })
                except (ValueError, TypeError) as e:
                    errors.append(f'第{i}行({name}): 数据格式错误 - {str(e)}')
                    skipped += 1

            # ---- 所有行处理完毕：按供应商分组生成采购订单 ----
            from collections import defaultdict
            supplier_groups = defaultdict(list)
            for item in processed_items:
                supplier_groups[item['supplier_name']].append(item)

            po_count = 0
            default_supplier = Supplier.query.first()

            for supplier_name, items in supplier_groups.items():
                # 匹配供应商：优先精确匹配，否则用默认供应商
                supplier = Supplier.query.filter_by(name=supplier_name).first()
                if not supplier and supplier_name:
                    # 模糊匹配
                    supplier = Supplier.query.filter(
                        Supplier.name.like(f'%{supplier_name}%')
                    ).first()
                if not supplier:
                    supplier = default_supplier
                if not supplier:
                    # 数据库里一个供应商都没有，跳过
                    continue

                # 计算总金额
                total_amount = 0
                for item_data in items:
                    total_amount += item_data['quantity'] * item_data['unit_price']

                # 1. 创建采购订单（草稿）
                order = PurchaseOrder(
                    order_no=generate_order_no('PO'),
                    supplier_id=supplier.id,
                    status='draft',
                    total_amount=total_amount,
                    remark=f'【批量导入】{supplier_name or "未指定供应商"}'
                )
                db.session.add(order)
                db.session.flush()

                # 创建订单明细
                for item_data in items:
                    p = item_data['product']
                    _qty = item_data['quantity']
                    _up = item_data['unit_price']
                    order_item = PurchaseOrderItem(
                        order_id=order.id,
                        product_id=p.id,
                        quantity=_qty,
                        unit_price=_up,
                        subtotal=_qty * _up
                    )
                    db.session.add(order_item)

                # 导入订单保持草稿，由人工操作审批→付款→入库
                po_count += 1

        result_msg = f'导入完成：新增 {created} 条，更新 {updated} 条，跳过 {skipped} 条'
        if po_count > 0:
            result_msg += f'，生成草稿采购订单 {po_count} 张，请前往采购进货页面操作审批→付款→入库'

        return jsonify({
            'message': result_msg,
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:20],
            'products': created_products[:50],
            'purchase_orders': po_count
        })

    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400


def _parse_csv(content):
    """解析CSV内容"""
    # 尝试检测编码（utf-8-sig 优先，可自动去除BOM）
    for encoding in ['utf-8-sig', 'utf-8', 'gbk', 'gb2312']:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode('utf-8', errors='replace')

    reader = csv.DictReader(io.StringIO(text))
    # 清理列名：去除BOM、前后空格（WPS/Excel保存的CSV可能携带）
    raw_rows = list(reader)
    clean_rows = []
    for row in raw_rows:
        clean_row = {}
        for key, val in row.items():
            ck = key.replace('\ufeff', '').strip()  # 去除BOM和空格
            clean_row[ck] = val
        clean_rows.append(clean_row)
    return clean_rows


def _parse_xlsx(content):
    """解析Excel文件"""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    rows = []
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # 清理列名：去除BOM和空格
    headers = [str(h).replace('\ufeff', '').strip() if h else '' for h in raw_headers]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = str(value) if value is not None else ''
        if row_dict:
            rows.append(row_dict)

    return rows


@app.route('/api/products/template', methods=['GET'])
def download_template():
    """下载商品导入模板（CSV格式）"""
    csv_content = (
        '商品编码,商品名称,类别,型号,供应商,单位,采购价,销售价,安全库存,库存数量\n'
        'LX-20-S,20寸登机箱-银色,行李箱,LX-20,新秀丽,个,180,359,50,50\n'
        'LX-24-S,24寸托运箱-银色,行李箱,LX-24,新秀丽,个,260,499,40,40\n'
        'LX-28-T,28寸托运箱-钛金灰,行李箱,LX-28,外交官,个,320,659,30,30\n'
    )
    # 添加BOM以支持Excel正确显示中文
    bom = '\ufeff'
    resp = Response(bom + csv_content, mimetype='text/csv')
    resp.headers['Content-Disposition'] = 'attachment; filename=product_template.csv'
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
    return resp


# ============================================================
# 路由 - 供应商API
# ============================================================

@app.route('/api/suppliers', methods=['GET'])
def list_suppliers():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    query = Supplier.query.order_by(Supplier.id.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/suppliers', methods=['POST'])
def create_supplier():
    data = request.json
    supplier = Supplier(
        name=data['name'], contact_person=data.get('contact_person', ''),
        phone=data.get('phone', ''), email=data.get('email', ''),
        address=data.get('address', ''),
        rating=data.get('rating', 3),
        cooperation_history=data.get('cooperation_history', '')
    )
    db.session.add(supplier)
    db.session.commit()
    return jsonify(supplier.to_dict()), 201


@app.route('/api/suppliers/<int:sid>', methods=['PUT'])
def update_supplier(sid):
    """更新供应商"""
    supplier = Supplier.query.get(sid)
    if not supplier:
        return jsonify({'error': '供应商不存在'}), 404
    data = request.json
    for key in ['name', 'contact_person', 'phone', 'email', 'address', 'rating', 'cooperation_history']:
        if key in data:
            setattr(supplier, key, data[key])
    db.session.commit()
    return jsonify(supplier.to_dict())


@app.route('/api/suppliers/<int:sid>', methods=['DELETE'])
def delete_supplier(sid):
    """删除供应商"""
    supplier = Supplier.query.get(sid)
    if not supplier:
        return jsonify({'error': '供应商不存在'}), 404

    try:
        with transaction():
            # 先删除关联的采购订单明细和订单
            orders = PurchaseOrder.query.filter_by(supplier_id=sid).all()
            for order in orders:
                PurchaseOrderItem.query.filter_by(order_id=order.id).delete()
                db.session.delete(order)
            db.session.delete(supplier)
        return jsonify({'message': '删除成功'})
    except Exception as e:
        return jsonify({'error': f'删除失败: {str(e)}'}), 400


# ============================================================
# 路由 - 客户API
# ============================================================

@app.route('/api/customers', methods=['GET'])
def list_customers():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    query = Customer.query.order_by(Customer.id.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/customers', methods=['POST'])
def create_customer():
    data = request.json
    customer = Customer(
        name=data['name'], contact_person=data.get('contact_person', ''),
        phone=data.get('phone', ''), email=data.get('email', ''),
        address=data.get('address', ''),
        region=data.get('region', '')
    )
    db.session.add(customer)
    db.session.commit()
    return jsonify(customer.to_dict()), 201


@app.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    """更新客户"""
    customer = Customer.query.get(cid)
    if not customer:
        return jsonify({'error': '客户不存在'}), 404
    data = request.json
    for key in ['name', 'contact_person', 'phone', 'email', 'address', 'region']:
        if key in data:
            setattr(customer, key, data[key])
    db.session.commit()
    return jsonify(customer.to_dict())


@app.route('/api/customers/<int:cid>', methods=['DELETE'])
def delete_customer(cid):
    """删除客户"""
    customer = Customer.query.get(cid)
    if not customer:
        return jsonify({'error': '客户不存在'}), 404

    try:
        with transaction():
            # 先删除关联的销售订单明细和订单
            orders = SalesOrder.query.filter_by(customer_id=cid).all()
            for order in orders:
                SalesOrderItem.query.filter_by(order_id=order.id).delete()
                db.session.delete(order)
            db.session.delete(customer)
        return jsonify({'message': '删除成功'})
    except Exception as e:
        return jsonify({'error': f'删除失败: {str(e)}'}), 400


# ============================================================
# 路由 - 进货管理API (核心功能1)
# ============================================================

@app.route('/api/purchase-orders', methods=['GET'])
def list_purchase_orders():
    """采购订单列表（支持状态筛选）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    query = PurchaseOrder.query
    if status:
        query = query.filter(PurchaseOrder.status == status)
    query = query.order_by(PurchaseOrder.created_at.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/purchase-orders', methods=['POST'])
def create_purchase_order():
    """创建采购订单"""
    data = request.json
    try:
        with transaction():
            order = PurchaseOrder(
                order_no=generate_order_no('PO'),
                supplier_id=data['supplier_id'],
                status='draft',
                remark=data.get('remark', '')
            )
            db.session.add(order)
            db.session.flush()

            total = 0
            for item_data in data.get('items', []):
                product = Product.query.get(item_data['product_id'])
                if not product:
                    raise ValueError(f'商品ID {item_data["product_id"]} 不存在')

                unit_price = item_data.get('unit_price', product.purchase_price)
                qty = item_data['quantity']
                subtotal = qty * unit_price
                total += subtotal

                order_item = PurchaseOrderItem(
                    order_id=order.id,
                    product_id=item_data['product_id'],
                    quantity=qty,
                    unit_price=unit_price,
                    subtotal=subtotal
                )
                db.session.add(order_item)

            order.total_amount = total

        return jsonify(order.to_dict()), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/purchase-orders/<int:oid>/receive', methods=['POST'])
def receive_purchase_order(oid):
    """
    核心：采购入库
    确认收货 → 库存自动增加
    """
    order = PurchaseOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'paid':
        return jsonify({'error': f'订单状态为"{order.status}"，需先付款后才可入库'}), 400

    try:
        with transaction():
            for item in order.items:
                ok, msg = update_inventory(
                    product_id=item.product_id,
                    change_qty=item.quantity,
                    change_type='purchase_in',
                    ref_no=order.order_no,
                    remark=f'采购入库: {order.order_no}'
                )
                if not ok:
                    raise ValueError(msg)

            order.status = 'received'
        return jsonify({'message': '入库成功', 'order': order.to_dict()})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@app.route('/api/purchase-orders/<int:oid>/approve', methods=['POST'])
def approve_purchase_order(oid):
    """审批采购订单"""
    order = PurchaseOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'draft':
        return jsonify({'error': '只能审批草稿状态的订单'}), 400

    order.status = 'approved'
    db.session.commit()
    return jsonify({'message': '审批成功', 'order': order.to_dict()})


@app.route('/api/purchase-orders/<int:oid>/pay', methods=['POST'])
def pay_purchase_order(oid):
    """付款确认"""
    order = PurchaseOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'approved':
        return jsonify({'error': '只能对已审批的订单进行付款'}), 400

    order.status = 'paid'
    db.session.commit()
    return jsonify({'message': '付款成功', 'order': order.to_dict()})

@app.route('/api/purchase-orders/batch-delete', methods=['POST'])
def batch_delete_purchase_orders():
    """批量删除采购订单（仅草稿/已取消可删除）"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的订单ID'}), 400
    deleted = 0
    for oid in ids:
        order = PurchaseOrder.query.get(oid)
        if order and order.status in ('draft', 'cancelled'):
            for item in order.items:
                db.session.delete(item)
            db.session.delete(order)
            deleted += 1
    db.session.commit()
    return jsonify({'message': f'成功删除{deleted}个订单', 'deleted': deleted})


# ============================================================
# 路由 - 销售管理API (核心功能2)
# ============================================================

@app.route('/api/sales-orders', methods=['GET'])
def list_sales_orders():
    """销售订单列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    query = SalesOrder.query
    if status:
        query = query.filter(SalesOrder.status == status)
    query = query.order_by(SalesOrder.created_at.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/sales-orders', methods=['POST'])
def create_sales_order():
    """创建销售订单"""
    data = request.json
    try:
        with transaction():
            order = SalesOrder(
                order_no=generate_order_no('SO'),
                customer_id=data['customer_id'],
                status='draft',
                remark=data.get('remark', '')
            )
            db.session.add(order)
            db.session.flush()

            total = 0
            for item_data in data.get('items', []):
                product = Product.query.get(item_data['product_id'])
                if not product:
                    raise ValueError(f'商品ID {item_data["product_id"]} 不存在')

                unit_price = item_data.get('unit_price', product.sale_price)
                qty = item_data['quantity']
                subtotal = qty * unit_price
                total += subtotal

                order_item = SalesOrderItem(
                    order_id=order.id,
                    product_id=item_data['product_id'],
                    quantity=qty,
                    unit_price=unit_price,
                    subtotal=subtotal
                )
                db.session.add(order_item)

            order.total_amount = total

        return jsonify(order.to_dict()), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/sales-orders/<int:oid>/confirm', methods=['POST'])
def confirm_sales_order(oid):
    """确认销售订单"""
    order = SalesOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'draft':
        return jsonify({'error': '只能确认草稿状态的订单'}), 400
    order.status = 'confirmed'
    db.session.commit()
    return jsonify({'message': '确认成功', 'order': order.to_dict()})


@app.route('/api/sales-orders/<int:oid>/collect', methods=['POST'])
def collect_sales_order(oid):
    """收款确认"""
    order = SalesOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'confirmed':
        return jsonify({'error': '只能对已确认的订单进行收款'}), 400
    order.status = 'collected'
    db.session.commit()
    return jsonify({'message': '收款成功', 'order': order.to_dict()})


@app.route('/api/sales-orders/<int:oid>/ship', methods=['POST'])
def ship_sales_order(oid):
    """
    核心：销售出库
    发货 → 库存自动减少
    """
    order = SalesOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'collected':
        return jsonify({'error': f'订单状态为"{order.status}"，需先收款后才可发货'}), 400

    try:
        with transaction():
            for item in order.items:
                ok, msg = update_inventory(
                    product_id=item.product_id,
                    change_qty=-item.quantity,
                    change_type='sales_out',
                    ref_no=order.order_no,
                    remark=f'销售出库: {order.order_no}'
                )
                if not ok:
                    raise ValueError(msg)

            order.status = 'shipped'
        return jsonify({'message': '发货成功', 'order': order.to_dict()})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


@app.route('/api/sales-orders/<int:oid>/complete', methods=['POST'])
def complete_sales_order(oid):
    """完成销售订单"""
    order = SalesOrder.query.get(oid)
    if not order:
        return jsonify({'error': '订单不存在'}), 404
    if order.status != 'shipped':
        return jsonify({'error': '只能完成已发货的订单'}), 400
    order.status = 'completed'
    db.session.commit()
    return jsonify({'message': '完成成功', 'order': order.to_dict()})

@app.route('/api/sales-orders/batch-delete', methods=['POST'])
def batch_delete_sales_orders():
    """批量删除销售订单（仅草稿/已取消可删除）"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的订单ID'}), 400
    deleted = 0
    for oid in ids:
        order = SalesOrder.query.get(oid)
        if order and order.status in ('draft', 'cancelled'):
            for item in order.items:
                db.session.delete(item)
            db.session.delete(order)
            deleted += 1
    db.session.commit()
    return jsonify({'message': f'成功删除{deleted}个订单', 'deleted': deleted})


# ============================================================
# 路由 - 库存管理API (核心功能3)
# ============================================================

@app.route('/api/inventory', methods=['GET'])
def list_inventory():
    """实时库存列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '', type=str)
    query = Product.query
    if search:
        query = query.filter(
            or_(Product.code.contains(search), Product.name.contains(search),
                Product.manufacturer.contains(search))
        )
    query = query.order_by(Product.code)
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/inventory/returnable', methods=['GET'])
def returnable_inventory():
    """退货可用库存：有库存的商品 + 自动匹配供应商"""
    products = Product.query.filter(
        Product.id.in_(
            db.session.query(Inventory.product_id).filter(Inventory.quantity > 0)
        )
    ).order_by(Product.code).all()

    # 为每个商品查找供应商：优先 manufacturer 匹配，兜底采购历史
    result = []
    for p in products:
        inv = p.inventory
        supplier_id = None
        supplier_name = ''

        # 1. 优先：用manufacturer字段匹配供应商名称
        if p.manufacturer:
            fallback = Supplier.query.filter(Supplier.name == p.manufacturer).first()
            if fallback:
                supplier_id = fallback.id
                supplier_name = fallback.name

        # 2. 兜底：查最近一次采购记录
        if not supplier_id:
            latest_po = db.session.query(PurchaseOrder.supplier_id, Supplier.name)\
                .join(PurchaseOrderItem, PurchaseOrderItem.order_id == PurchaseOrder.id)\
                .join(Supplier, Supplier.id == PurchaseOrder.supplier_id)\
                .filter(PurchaseOrderItem.product_id == p.id)\
                .filter(PurchaseOrder.status == 'received')\
                .order_by(PurchaseOrder.created_at.desc())\
                .first()
            if latest_po:
                supplier_id = latest_po[0]
                supplier_name = latest_po[1]

        item = p.to_dict()
        item['supplier_id'] = supplier_id
        item['supplier_name'] = supplier_name
        result.append(item)

    return jsonify(result)


@app.route('/api/inventory/alerts', methods=['GET'])
def inventory_alerts():
    """库存预警列表"""
    alerts = get_stock_alerts()
    return jsonify({'count': len(alerts), 'alerts': alerts})


@app.route('/api/inventory/logs', methods=['GET'])
def inventory_logs():
    """库存变动日志"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    product_id = request.args.get('product_id', type=int)
    query = InventoryLog.query
    if product_id:
        query = query.filter(InventoryLog.product_id == product_id)
    query = query.order_by(InventoryLog.created_at.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/inventory/adjust', methods=['POST'])
def adjust_inventory():
    """库存盘点/调整"""
    data = request.json
    product_id = data.get('product_id')
    new_qty = data.get('quantity', 0)

    inv = Inventory.query.filter_by(product_id=product_id).first()
    if not inv:
        return jsonify({'error': '库存记录不存在'}), 404

    change_qty = new_qty - inv.quantity
    if change_qty == 0:
        return jsonify({'message': '库存无变化'})

    try:
        with transaction():
            ok, msg = update_inventory(
                product_id=product_id,
                change_qty=change_qty,
                change_type='adjust',
                ref_no=f'ADJ{datetime.now().strftime("%Y%m%d%H%M%S")}',
                remark=data.get('remark', '手动盘点调整')
            )
            if not ok:
                raise ValueError(msg)
        return jsonify({'message': '调整成功'})
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400


# ============================================================
# 路由 - 统计数据API
# ============================================================

@app.route('/api/stats/sales-rank', methods=['GET'])
def sales_rank():
    """商品销售排行"""
    results = db.session.query(
        Product.name,
        func.coalesce(func.sum(SalesOrderItem.quantity), 0).label('total_qty'),
        func.coalesce(func.sum(SalesOrderItem.subtotal), 0).label('total_amount')
    ).join(SalesOrderItem, Product.id == SalesOrderItem.product_id
    ).join(SalesOrder, SalesOrderItem.order_id == SalesOrder.id
    ).filter(SalesOrder.status.in_(['shipped', 'completed'])
    ).group_by(Product.id).order_by(text('total_qty DESC')).limit(10).all()

    return jsonify([{
        'product_name': name,
        'total_quantity': int(qty),
        'total_amount': round(amount, 2)
    } for name, qty, amount in results])


def _parse_date_arg(name):
    """解析 YYYY-MM-DD 日期参数，解析失败时返回 None。"""
    value = request.args.get(name, '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d')
    except ValueError:
        return None


def _date_range(default_days=30):
    """获取统计日期范围，结束日期按自然日闭区间处理。"""
    end_date = _parse_date_arg('end_date') or datetime.now()
    start_date = _parse_date_arg('start_date') or (end_date - timedelta(days=default_days - 1))
    start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)
    return start_date, end_date


def _add_months(dt, months):
    """按月偏移日期，只用于统计默认范围和补零。"""
    year = dt.year + (dt.month - 1 + months) // 12
    month = (dt.month - 1 + months) % 12 + 1
    return dt.replace(year=year, month=month, day=1)


def _profit_rate(profit, amount):
    """计算毛利率，销售额为 0 时返回 0。"""
    amount = amount or 0
    if amount == 0:
        return 0
    return round((profit or 0) / amount * 100, 2)


@app.route('/api/stats/sales-trend', methods=['GET'])
def sales_trend():
    """按日/月统计销售额折线图数据"""
    period = request.args.get('period', 'day')
    if period not in ('day', 'month'):
        period = 'day'

    if period == 'month':
        end_base = _parse_date_arg('end_date') or datetime.now()
        end_month = end_base.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        start_month = (_parse_date_arg('start_date') or _add_months(end_month, -11)).replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        )
        end_limit = _add_months(end_month, 1)
        bucket_expr = func.strftime('%Y-%m', SalesOrder.updated_at)
        current = start_month
        labels = []
        while current < end_limit:
            labels.append(current.strftime('%Y-%m'))
            current = _add_months(current, 1)
        start_date, end_date = start_month, end_limit
    else:
        start_date, end_date = _date_range(default_days=30)
        bucket_expr = func.strftime('%Y-%m-%d', SalesOrder.updated_at)
        labels = []
        current = start_date
        while current < end_date:
            labels.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)

    rows = db.session.query(
        bucket_expr.label('bucket'),
        func.coalesce(func.sum(SalesOrder.total_amount), 0).label('sales_amount'),
        func.count(SalesOrder.id).label('order_count')
    ).filter(
        SalesOrder.status.in_(['shipped', 'completed']),
        SalesOrder.updated_at >= start_date,
        SalesOrder.updated_at < end_date
    ).group_by('bucket').order_by('bucket').all()

    row_map = {bucket: {'sales_amount': float(amount or 0), 'order_count': int(count or 0)}
               for bucket, amount, count in rows}
    data = [{
        'label': label,
        'sales_amount': round(row_map.get(label, {}).get('sales_amount', 0), 2),
        'order_count': row_map.get(label, {}).get('order_count', 0)
    } for label in labels]

    return jsonify({
        'period': period,
        'start_date': labels[0] if labels else '',
        'end_date': labels[-1] if labels else '',
        'data': data
    })


@app.route('/api/stats/profit-products', methods=['GET'])
def profit_products():
    """按商品统计销售利润"""
    limit = request.args.get('limit', 10, type=int)
    limit = max(1, min(limit, 50))
    start_date, end_date = _date_range(default_days=90)

    sales_amount_expr = func.coalesce(func.sum(SalesOrderItem.quantity * SalesOrderItem.unit_price), 0)
    cost_amount_expr = func.coalesce(func.sum(SalesOrderItem.quantity * Product.purchase_price), 0)
    qty_expr = func.coalesce(func.sum(SalesOrderItem.quantity), 0)

    rows = db.session.query(
        Product.id,
        Product.name,
        qty_expr.label('total_quantity'),
        sales_amount_expr.label('sales_amount'),
        cost_amount_expr.label('cost_amount'),
        (sales_amount_expr - cost_amount_expr).label('profit_amount')
    ).join(SalesOrderItem, Product.id == SalesOrderItem.product_id
    ).join(SalesOrder, SalesOrderItem.order_id == SalesOrder.id
    ).filter(
        SalesOrder.status.in_(['shipped', 'completed']),
        SalesOrder.updated_at >= start_date,
        SalesOrder.updated_at < end_date
    ).group_by(Product.id).order_by(text('profit_amount DESC')).limit(limit).all()

    return jsonify([{
        'product_id': product_id,
        'product_name': name,
        'total_quantity': int(total_quantity or 0),
        'sales_amount': round(sales_amount or 0, 2),
        'cost_amount': round(cost_amount or 0, 2),
        'profit_amount': round(profit_amount or 0, 2),
        'profit_rate': _profit_rate(profit_amount, sales_amount)
    } for product_id, name, total_quantity, sales_amount, cost_amount, profit_amount in rows])


@app.route('/api/stats/profit-orders', methods=['GET'])
def profit_orders():
    """按销售订单统计利润"""
    limit = request.args.get('limit', 10, type=int)
    limit = max(1, min(limit, 50))
    start_date, end_date = _date_range(default_days=90)

    sales_amount_expr = func.coalesce(func.sum(SalesOrderItem.quantity * SalesOrderItem.unit_price), 0)
    cost_amount_expr = func.coalesce(func.sum(SalesOrderItem.quantity * Product.purchase_price), 0)

    rows = db.session.query(
        SalesOrder.id,
        SalesOrder.order_no,
        Customer.name,
        SalesOrder.status,
        SalesOrder.updated_at,
        sales_amount_expr.label('sales_amount'),
        cost_amount_expr.label('cost_amount'),
        (sales_amount_expr - cost_amount_expr).label('profit_amount')
    ).join(Customer, SalesOrder.customer_id == Customer.id
    ).join(SalesOrderItem, SalesOrderItem.order_id == SalesOrder.id
    ).join(Product, Product.id == SalesOrderItem.product_id
    ).filter(
        SalesOrder.status.in_(['shipped', 'completed']),
        SalesOrder.updated_at >= start_date,
        SalesOrder.updated_at < end_date
    ).group_by(SalesOrder.id).order_by(SalesOrder.updated_at.desc()).limit(limit).all()

    return jsonify([{
        'order_id': order_id,
        'order_no': order_no,
        'customer_name': customer_name,
        'status': status,
        'sales_amount': round(sales_amount or 0, 2),
        'cost_amount': round(cost_amount or 0, 2),
        'profit_amount': round(profit_amount or 0, 2),
        'profit_rate': _profit_rate(profit_amount, sales_amount),
        'updated_at': updated_at.strftime('%Y-%m-%d %H:%M') if updated_at else ''
    } for order_id, order_no, customer_name, status, updated_at,
         sales_amount, cost_amount, profit_amount in rows])


@app.route('/api/reconciliation/customers', methods=['GET'])
def customer_reconciliation():
    """客户对账汇总：累计消费、已收款、欠款"""
    limit = request.args.get('limit', 10, type=int)
    limit = max(1, min(limit, 100))

    rows = db.session.query(
        Customer.id,
        Customer.name,
        func.coalesce(func.sum(case(
            (SalesOrder.status.in_(['shipped', 'completed']), SalesOrder.total_amount),
            else_=0
        )), 0).label('total_consumption'),
        func.coalesce(func.sum(case(
            (SalesOrder.status.in_(['collected', 'shipped', 'completed']), SalesOrder.total_amount),
            else_=0
        )), 0).label('collected_amount'),
        func.coalesce(func.sum(case(
            (SalesOrder.status == 'confirmed', SalesOrder.total_amount),
            else_=0
        )), 0).label('receivable_amount'),
        func.count(SalesOrder.id).label('order_count')
    ).outerjoin(SalesOrder, Customer.id == SalesOrder.customer_id
    ).group_by(Customer.id).order_by(text('total_consumption DESC')).limit(limit).all()

    return jsonify([{
        'customer_id': customer_id,
        'customer_name': name,
        'total_consumption': round(total_consumption or 0, 2),
        'collected_amount': round(collected_amount or 0, 2),
        'receivable_amount': round(receivable_amount or 0, 2),
        'order_count': int(order_count or 0)
    } for customer_id, name, total_consumption, collected_amount, receivable_amount, order_count in rows])


@app.route('/api/reconciliation/suppliers', methods=['GET'])
def supplier_reconciliation():
    """供应商对账汇总：累计采购额、已付款、待付款"""
    limit = request.args.get('limit', 10, type=int)
    limit = max(1, min(limit, 100))

    rows = db.session.query(
        Supplier.id,
        Supplier.name,
        func.coalesce(func.sum(case(
            (PurchaseOrder.status == 'received', PurchaseOrder.total_amount),
            else_=0
        )), 0).label('total_purchase'),
        func.coalesce(func.sum(case(
            (PurchaseOrder.status.in_(['paid', 'received']), PurchaseOrder.total_amount),
            else_=0
        )), 0).label('paid_amount'),
        func.coalesce(func.sum(case(
            (PurchaseOrder.status == 'approved', PurchaseOrder.total_amount),
            else_=0
        )), 0).label('payable_amount'),
        func.count(PurchaseOrder.id).label('order_count')
    ).outerjoin(PurchaseOrder, Supplier.id == PurchaseOrder.supplier_id
    ).group_by(Supplier.id).order_by(text('total_purchase DESC')).limit(limit).all()

    return jsonify([{
        'supplier_id': supplier_id,
        'supplier_name': name,
        'total_purchase': round(total_purchase or 0, 2),
        'paid_amount': round(paid_amount or 0, 2),
        'payable_amount': round(payable_amount or 0, 2),
        'order_count': int(order_count or 0)
    } for supplier_id, name, total_purchase, paid_amount, payable_amount, order_count in rows])


# ============================================================
# 启动入口
# ============================================================

def init_demo_data():
    """初始化演示数据"""
    if Product.query.first():
        return  # 已有数据则跳过

    # 供应商
    s1 = Supplier(name='华为技术有限公司', contact_person='张经理', phone='13800001001',
                  email='zhang@huawei.com', address='深圳市龙岗区', rating=5,
                  cooperation_history='长期战略合作伙伴')
    s2 = Supplier(name='小米科技有限公司', contact_person='李主管', phone='13800001002',
                  email='li@xiaomi.com', address='北京市海淀区', rating=4)
    s3 = Supplier(name='京东供应链', contact_person='王经理', phone='13800001003',
                  email='wang@jd.com', address='北京市亦庄', rating=4)
    db.session.add_all([s1, s2, s3])

    # 客户
    c1 = Customer(name='北京朝阳科技城', contact_person='赵总', phone='13900002001',
                  email='zhao@chaoyang.com', address='北京市朝阳区', region='华北')
    c2 = Customer(name='上海浦东电子市场', contact_person='钱经理', phone='13900002002',
                  email='qian@pudong.com', address='上海市浦东新区', region='华东')
    c3 = Customer(name='广州天河数码广场', contact_person='孙老板', phone='13900002003',
                  email='sun@tianhe.com', address='广州市天河区', region='华南')
    db.session.add_all([c1, c2, c3])

    # 商品
    products_data = [
        {'code': 'P001', 'name': '华为Mate 60 Pro', 'category': '手机', 'unit': '台',
         'purchase_price': 5500, 'sale_price': 6999, 'safety_stock': 20,
         'manufacturer': '华为技术有限公司'},
        {'code': 'P002', 'name': '小米14 Ultra', 'category': '手机', 'unit': '台',
         'purchase_price': 4200, 'sale_price': 5999, 'safety_stock': 15,
         'manufacturer': '小米科技有限公司'},
        {'code': 'P003', 'name': '联想ThinkPad X1', 'category': '笔记本电脑', 'unit': '台',
         'purchase_price': 7200, 'sale_price': 9999, 'safety_stock': 10,
         'manufacturer': '京东供应链'},
        {'code': 'P004', 'name': 'AirPods Pro 2', 'category': '耳机', 'unit': '个',
         'purchase_price': 1400, 'sale_price': 1899, 'safety_stock': 30,
         'manufacturer': '京东供应链'},
        {'code': 'P005', 'name': 'iPad Air 5', 'category': '平板', 'unit': '台',
         'purchase_price': 3800, 'sale_price': 4799, 'safety_stock': 15,
         'manufacturer': '京东供应链'},
        {'code': 'P006', 'name': '索尼WH-1000XM5', 'category': '耳机', 'unit': '个',
         'purchase_price': 1800, 'sale_price': 2499, 'safety_stock': 20,
         'manufacturer': '京东供应链'},
        {'code': 'P007', 'name': '戴尔显示器U2723QE', 'category': '显示器', 'unit': '台',
         'purchase_price': 2800, 'sale_price': 3999, 'safety_stock': 12,
         'manufacturer': '京东供应链'},
        {'code': 'P008', 'name': '罗技MX Master 3S', 'category': '外设', 'unit': '个',
         'purchase_price': 450, 'sale_price': 699, 'safety_stock': 40,
         'manufacturer': '华为技术有限公司'},
    ]

    for pd_ in products_data:
        p = Product(**pd_)
        db.session.add(p)
        db.session.flush()
        inv = Inventory(product_id=p.id, quantity=0)
        db.session.add(inv)
        # 给部分商品一些初始库存
        if pd_['code'] in ['P001', 'P003', 'P006', 'P008']:
            inv.quantity = pd_['safety_stock'] * 3

    db.session.commit()

    # 初始库存日志
    for p in Product.query.all():
        inv = Inventory.query.filter_by(product_id=p.id).first()
        if inv and inv.quantity > 0:
            log = InventoryLog(
                product_id=p.id, change_type='adjust',
                change_quantity=inv.quantity, before_quantity=0,
                after_quantity=inv.quantity,
                reference_no='INIT',
                remark='系统初始库存'
            )
            db.session.add(log)
    db.session.commit()


# ============================================================
# 退货单管理
# ============================================================

@app.route('/api/return-orders', methods=['GET'])
def get_return_orders():
    """获取退货单列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    query = ReturnOrder.query
    if status:
        query = query.filter_by(status=status)
    query = query.order_by(ReturnOrder.created_at.desc())
    return jsonify(paginate_query(query, page, per_page))


@app.route('/api/return-orders', methods=['POST'])
def create_return_order():
    """创建退货单（草稿）"""
    data = request.json
    supplier_id = data.get('supplier_id')
    items_data = data.get('items', [])

    if not supplier_id:
        return jsonify({'error': '请选择供应商'}), 400
    if not items_data:
        return jsonify({'error': '请至少添加一个退货商品'}), 400

    try:
        with transaction():
            order = ReturnOrder(
                order_no=generate_order_no('RO'),
                supplier_id=supplier_id,
                status='draft',
                remark=data.get('remark', ''),
                total_amount=0
            )
            db.session.add(order)
            db.session.flush()

            total = 0
            for item in items_data:
                qty = int(item.get('quantity', 0))
                price = float(item.get('unit_price', 0))
                if qty <= 0:
                    raise ValueError('退货数量必须大于0')
                sub = qty * price
                total += sub
                ri = ReturnOrderItem(
                    order_id=order.id,
                    product_id=int(item['product_id']),
                    quantity=qty,
                    unit_price=price,
                    subtotal=sub
                )
                db.session.add(ri)

            order.total_amount = total

        return jsonify(order.to_dict()), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'创建失败: {str(e)}'}), 500


@app.route('/api/return-orders/<int:oid>/submit', methods=['POST'])
def submit_return_order(oid):
    """提交退货申请"""
    order = ReturnOrder.query.get_or_404(oid)
    if order.status != 'draft':
        return jsonify({'error': '只有草稿状态的退货单才能提交'}), 400
    order.status = 'submitted'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/return-orders/<int:oid>/approve', methods=['POST'])
def approve_return_order(oid):
    """审核通过退货单"""
    order = ReturnOrder.query.get_or_404(oid)
    if order.status != 'submitted':
        return jsonify({'error': '只有已申请状态的退货单才能审核'}), 400
    order.status = 'approved'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/return-orders/<int:oid>/return', methods=['POST'])
def return_goods(oid):
    """确认退货出库，扣减库存"""
    order = ReturnOrder.query.get_or_404(oid)
    if order.status != 'approved':
        return jsonify({'error': '只有已审核状态的退货单才能退货出库'}), 400

    try:
        with transaction():
            for item in order.items.all():
                inv = Inventory.query.filter_by(product_id=item.product_id).first()
                if not inv:
                    raise ValueError(f'商品 {item.product.name if item.product else "?"} 无库存记录')
                if inv.quantity < item.quantity:
                    raise ValueError(f'商品 {item.product.name if item.product else "?"} 库存不足 (当前:{inv.quantity}, 退货:{item.quantity})')

                ok, msg = update_inventory(
                    product_id=item.product_id,
                    change_qty=-item.quantity,
                    change_type='return_out',
                    ref_no=order.order_no,
                    remark=f'退货出库: {order.order_no}'
                )
                if not ok:
                    raise ValueError(msg)

            order.status = 'returned'
            order.updated_at = datetime.now()

        return jsonify(order.to_dict())
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'退货出库失败: {str(e)}'}), 500


@app.route('/api/return-orders/<int:oid>/refund', methods=['POST'])
def refund_return_order(oid):
    """确认退款"""
    order = ReturnOrder.query.get_or_404(oid)
    if order.status != 'returned':
        return jsonify({'error': '只有已退货状态的退货单才能退款'}), 400
    order.status = 'refunded'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/return-orders/<int:oid>/complete', methods=['POST'])
def complete_return_order(oid):
    """完成退货单"""
    order = ReturnOrder.query.get_or_404(oid)
    if order.status != 'refunded':
        return jsonify({'error': '只有已退款状态的退货单才能完成'}), 400
    order.status = 'completed'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/return-orders/<int:oid>/cancel', methods=['POST'])
def cancel_return_order(oid):
    """取消退货单"""
    order = ReturnOrder.query.get_or_404(oid)
    if order.status in ('completed', 'cancelled'):
        return jsonify({'error': '已完成/已取消的退货单无法取消'}), 400
    order.status = 'cancelled'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())

@app.route('/api/return-orders/batch-delete', methods=['POST'])
def batch_delete_return_orders():
    """批量删除退货单（仅草稿/已取消可删除）"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的订单ID'}), 400
    deleted = 0
    for oid in ids:
        order = ReturnOrder.query.get(oid)
        if order and order.status in ('draft', 'cancelled'):
            for item in order.items:
                db.session.delete(item)
            db.session.delete(order)
            deleted += 1
    db.session.commit()
    return jsonify({'message': f'成功删除{deleted}个订单', 'deleted': deleted})


# ============================================================
# 客户退货 API
# ============================================================

@app.route('/api/customer-returns', methods=['GET'])
def get_customer_returns():
    """获取客户退货单列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    q = CustomerReturnOrder.query.order_by(CustomerReturnOrder.created_at.desc())
    if status:
        q = q.filter(CustomerReturnOrder.status == status)
    return jsonify(paginate_query(q, page, per_page))


@app.route('/api/customer-returns', methods=['POST'])
def create_customer_return():
    """创建客户退货单（草稿）"""
    data = request.json
    customer_id = data.get('customer_id')
    items_data = data.get('items', [])

    if not customer_id:
        return jsonify({'error': '请选择客户'}), 400
    if not items_data:
        return jsonify({'error': '请至少添加一个退货商品'}), 400

    try:
        with transaction():
            order = CustomerReturnOrder(
                order_no=generate_order_no('CR'),
                customer_id=customer_id,
                status='draft',
                remark=data.get('remark', ''),
                total_amount=0
            )
            db.session.add(order)
            db.session.flush()

            total = 0
            for item in items_data:
                qty = int(item.get('quantity', 0))
                price = float(item.get('unit_price', 0))
                if qty <= 0:
                    raise ValueError('退货数量必须大于0')
                sub = qty * price
                total += sub
                ri = CustomerReturnOrderItem(
                    order_id=order.id,
                    product_id=int(item['product_id']),
                    quantity=qty,
                    unit_price=price,
                    subtotal=sub
                )
                db.session.add(ri)

            order.total_amount = total

        return jsonify(order.to_dict()), 201
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'创建失败: {str(e)}'}), 500


@app.route('/api/customer-returns/<int:oid>/submit', methods=['POST'])
def submit_customer_return(oid):
    """提交客户退货申请"""
    order = CustomerReturnOrder.query.get_or_404(oid)
    if order.status != 'draft':
        return jsonify({'error': '只有草稿状态才能提交'}), 400
    order.status = 'submitted'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/customer-returns/<int:oid>/approve', methods=['POST'])
def approve_customer_return(oid):
    """审核通过客户退货单"""
    order = CustomerReturnOrder.query.get_or_404(oid)
    if order.status != 'submitted':
        return jsonify({'error': '只有已申请的退货单才能审核'}), 400
    order.status = 'approved'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/customer-returns/<int:oid>/receive', methods=['POST'])
def receive_customer_return(oid):
    """确认收货（客户退货入库，回补库存）"""
    order = CustomerReturnOrder.query.get_or_404(oid)
    if order.status != 'approved':
        return jsonify({'error': '只有已审核的退货单才能收货'}), 400
    with transaction():
        for item in order.items.all():
            update_inventory(item.product_id, item.quantity, 'sales_return_in',
                             f'客户退货入库 CR-{order.order_no}')
        order.status = 'received'
        order.updated_at = datetime.now()
    return jsonify(order.to_dict())


@app.route('/api/customer-returns/<int:oid>/refund', methods=['POST'])
def refund_customer_return(oid):
    """退款给客户"""
    order = CustomerReturnOrder.query.get_or_404(oid)
    if order.status != 'received':
        return jsonify({'error': '只有已收货的退货单才能退款'}), 400
    order.status = 'refunded'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/customer-returns/<int:oid>/complete', methods=['POST'])
def complete_customer_return(oid):
    """完成客户退货单"""
    order = CustomerReturnOrder.query.get_or_404(oid)
    if order.status != 'refunded':
        return jsonify({'error': '只有已退款的退货单才能完成'}), 400
    order.status = 'completed'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())


@app.route('/api/customer-returns/<int:oid>/cancel', methods=['POST'])
def cancel_customer_return(oid):
    """取消客户退货单"""
    order = CustomerReturnOrder.query.get_or_404(oid)
    if order.status in ('completed', 'cancelled'):
        return jsonify({'error': '已完成/已取消的退货单无法取消'}), 400
    order.status = 'cancelled'
    order.updated_at = datetime.now()
    db.session.commit()
    return jsonify(order.to_dict())

@app.route('/api/customer-returns/batch-delete', methods=['POST'])
def batch_delete_customer_returns():
    """批量删除客户退货单（仅草稿/已取消可删除）"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的订单ID'}), 400
    deleted = 0
    for oid in ids:
        order = CustomerReturnOrder.query.get(oid)
        if order and order.status in ('draft', 'cancelled'):
            for item in order.items:
                db.session.delete(item)
            db.session.delete(order)
            deleted += 1
    db.session.commit()
    return jsonify({'message': f'成功删除{deleted}个订单', 'deleted': deleted})


# ============================================================
# 数据库模型 - 商品报损/报溢
# ============================================================

class DamageReport(db.Model):
    """商品报损单"""
    __tablename__ = 'damage_reports'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    report_no = db.Column(db.String(30), unique=True, nullable=False, comment='报损单号')
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, comment='报损数量')
    reason = db.Column(db.String(50), nullable=False, comment='报损原因')
    remark = db.Column(db.String(200), comment='备注')
    status = db.Column(db.String(20), default='draft', comment='draft/submitted/approved/executed/cancelled')
    loss_amount = db.Column(db.Float, default=0, comment='损失金额')
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    product = db.relationship('Product')

    def to_dict(self):
        return {
            'id': self.id, 'report_no': self.report_no,
            'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'product_code': self.product.code if self.product else '',
            'quantity': self.quantity, 'reason': self.reason,
            'remark': self.remark, 'status': self.status,
            'loss_amount': self.loss_amount,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else '',
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M') if self.updated_at else ''
        }


class OverflowReport(db.Model):
    """商品报溢单"""
    __tablename__ = 'overflow_reports'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    report_no = db.Column(db.String(30), unique=True, nullable=False, comment='报溢单号')
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, comment='报溢数量')
    reason = db.Column(db.String(50), nullable=False, comment='报溢原因')
    remark = db.Column(db.String(200), comment='备注')
    status = db.Column(db.String(20), default='draft', comment='draft/submitted/approved/executed/cancelled')
    amount = db.Column(db.Float, default=0, comment='金额')
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    product = db.relationship('Product')

    def to_dict(self):
        return {
            'id': self.id, 'report_no': self.report_no,
            'product_id': self.product_id,
            'product_name': self.product.name if self.product else '',
            'product_code': self.product.code if self.product else '',
            'quantity': self.quantity, 'reason': self.reason,
            'remark': self.remark, 'status': self.status,
            'amount': self.amount,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M') if self.created_at else '',
            'updated_at': self.updated_at.strftime('%Y-%m-%d %H:%M') if self.updated_at else ''
        }


# ============================================================
# 商品报损 API
# ============================================================

@app.route('/api/damage-reports', methods=['GET'])
def get_damage_reports():
    """获取商品报损单列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    q = DamageReport.query.order_by(DamageReport.created_at.desc())
    if status:
        q = q.filter(DamageReport.status == status)
    return jsonify(paginate_query(q, page, per_page))


@app.route('/api/damage-reports', methods=['POST'])
def create_damage_report():
    """创建商品报损单（草稿）"""
    data = request.json
    product_id = data.get('product_id')
    quantity = int(data.get('quantity', 0))
    reason = data.get('reason', '')

    if not product_id:
        return jsonify({'error': '请选择商品'}), 400
    if quantity <= 0:
        return jsonify({'error': '报损数量必须大于0'}), 400
    if not reason:
        return jsonify({'error': '请选择报损原因'}), 400

    product = Product.query.get(product_id)
    if not product:
        return jsonify({'error': '商品不存在'}), 404

    inv = Inventory.query.filter_by(product_id=product_id).first()
    stock = inv.quantity if inv else 0
    if quantity > stock:
        return jsonify({'error': f'报损数量({quantity})不能大于当前库存({stock})'}), 400

    report = DamageReport(
        report_no=generate_order_no('DM'),
        product_id=product_id,
        quantity=quantity,
        reason=reason,
        remark=data.get('remark', ''),
        status='draft',
        loss_amount=round(quantity * product.purchase_price, 2)
    )
    db.session.add(report)
    db.session.commit()
    return jsonify(report.to_dict()), 201


@app.route('/api/damage-reports/<int:rid>/submit', methods=['POST'])
def submit_damage_report(rid):
    """提交报损单"""
    report = DamageReport.query.get_or_404(rid)
    if report.status != 'draft':
        return jsonify({'error': '只有草稿状态才能提交'}), 400
    report.status = 'submitted'
    report.updated_at = datetime.now()
    db.session.commit()
    return jsonify(report.to_dict())


@app.route('/api/damage-reports/<int:rid>/approve', methods=['POST'])
def approve_damage_report(rid):
    """审核通过报损单"""
    report = DamageReport.query.get_or_404(rid)
    if report.status != 'submitted':
        return jsonify({'error': '只有已提交的报损单才能审核'}), 400
    report.status = 'approved'
    report.updated_at = datetime.now()
    db.session.commit()
    return jsonify(report.to_dict())


@app.route('/api/damage-reports/<int:rid>/execute', methods=['POST'])
def execute_damage_report(rid):
    """执行报损（扣减库存）"""
    report = DamageReport.query.get_or_404(rid)
    if report.status != 'approved':
        return jsonify({'error': '只有已审核的报损单才能执行'}), 400

    with transaction():
        ok, msg = update_inventory(
            report.product_id,
            -report.quantity,
            'damage_out',
            report.report_no,
            f'商品报损: {report.reason}'
        )
        if not ok:
            raise ValueError(msg)
        report.status = 'executed'
        report.updated_at = datetime.now()

    return jsonify(report.to_dict())


@app.route('/api/damage-reports/<int:rid>/cancel', methods=['POST'])
def cancel_damage_report(rid):
    """取消报损单"""
    report = DamageReport.query.get_or_404(rid)
    if report.status in ('executed', 'cancelled'):
        return jsonify({'error': '已执行/已取消的报损单无法取消'}), 400
    report.status = 'cancelled'
    report.updated_at = datetime.now()
    db.session.commit()
    return jsonify(report.to_dict())


@app.route('/api/damage-reports/batch-delete', methods=['POST'])
def batch_delete_damage_reports():
    """批量删除报损单（仅草稿/已取消可删除）"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的ID'}), 400
    deleted = 0
    for rid in ids:
        report = DamageReport.query.get(rid)
        if report and report.status in ('draft', 'cancelled'):
            db.session.delete(report)
            deleted += 1
    db.session.commit()
    return jsonify({'message': f'成功删除{deleted}个报损单', 'deleted': deleted})


# ============================================================
# 商品报溢 API
# ============================================================

@app.route('/api/overflow-reports', methods=['GET'])
def get_overflow_reports():
    """获取商品报溢单列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    q = OverflowReport.query.order_by(OverflowReport.created_at.desc())
    if status:
        q = q.filter(OverflowReport.status == status)
    return jsonify(paginate_query(q, page, per_page))


@app.route('/api/overflow-reports', methods=['POST'])
def create_overflow_report():
    """创建商品报溢单（草稿）"""
    data = request.json
    product_id = data.get('product_id')
    quantity = int(data.get('quantity', 0))
    reason = data.get('reason', '')

    if not product_id:
        return jsonify({'error': '请选择商品'}), 400
    if quantity <= 0:
        return jsonify({'error': '报溢数量必须大于0'}), 400
    if not reason:
        return jsonify({'error': '请选择报溢原因'}), 400

    product = Product.query.get(product_id)
    if not product:
        return jsonify({'error': '商品不存在'}), 404

    report = OverflowReport(
        report_no=generate_order_no('OV'),
        product_id=product_id,
        quantity=quantity,
        reason=reason,
        remark=data.get('remark', ''),
        status='draft',
        amount=round(quantity * product.purchase_price, 2)
    )
    db.session.add(report)
    db.session.commit()
    return jsonify(report.to_dict()), 201


@app.route('/api/overflow-reports/<int:rid>/submit', methods=['POST'])
def submit_overflow_report(rid):
    """提交报溢单"""
    report = OverflowReport.query.get_or_404(rid)
    if report.status != 'draft':
        return jsonify({'error': '只有草稿状态才能提交'}), 400
    report.status = 'submitted'
    report.updated_at = datetime.now()
    db.session.commit()
    return jsonify(report.to_dict())


@app.route('/api/overflow-reports/<int:rid>/approve', methods=['POST'])
def approve_overflow_report(rid):
    """审核通过报溢单"""
    report = OverflowReport.query.get_or_404(rid)
    if report.status != 'submitted':
        return jsonify({'error': '只有已提交的报溢单才能审核'}), 400
    report.status = 'approved'
    report.updated_at = datetime.now()
    db.session.commit()
    return jsonify(report.to_dict())


@app.route('/api/overflow-reports/<int:rid>/execute', methods=['POST'])
def execute_overflow_report(rid):
    """执行报溢（减少库存）"""
    report = OverflowReport.query.get_or_404(rid)
    if report.status != 'approved':
        return jsonify({'error': '只有已审核的报溢单才能执行'}), 400

    with transaction():
        ok, msg = update_inventory(
            report.product_id,
            -report.quantity,
            'overflow_out',
            report.report_no,
            f'商品报溢: {report.reason}'
        )
        if not ok:
            raise ValueError(msg)
        report.status = 'executed'
        report.updated_at = datetime.now()

    return jsonify(report.to_dict())


@app.route('/api/overflow-reports/<int:rid>/cancel', methods=['POST'])
def cancel_overflow_report(rid):
    """取消报溢单"""
    report = OverflowReport.query.get_or_404(rid)
    if report.status in ('executed', 'cancelled'):
        return jsonify({'error': '已执行/已取消的报溢单无法取消'}), 400
    report.status = 'cancelled'
    report.updated_at = datetime.now()
    db.session.commit()
    return jsonify(report.to_dict())


@app.route('/api/overflow-reports/batch-delete', methods=['POST'])
def batch_delete_overflow_reports():
    """批量删除报溢单（仅草稿/已取消可删除）"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': '请提供要删除的ID'}), 400
    deleted = 0
    for rid in ids:
        report = OverflowReport.query.get(rid)
        if report and report.status in ('draft', 'cancelled'):
            db.session.delete(report)
            deleted += 1
    db.session.commit()
    return jsonify({'message': f'成功删除{deleted}个报溢单', 'deleted': deleted})

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        init_demo_data()
    print('=' * 60)
    print('  企业进销存智能管理系统 后端服务启动')
    print('  访问地址: http://localhost:5000')
    print('=' * 60)
    app.run(host='0.0.0.0', port=5000, debug=True)
